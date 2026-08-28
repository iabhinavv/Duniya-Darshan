#!/usr/bin/env python3
"""Read both budget workbooks and write the seed used by Duniya Darshan.

    python3 import_xlsx.py

World_tour_budget.xlsx  ->  the "World Tour" trip, one place per country.
India_tour_budget.xlsx  ->  the "Within India" trip, one place per STATE, with
                            each day's city kept as an itinerary entry. The India
                            sheet budgets by state and the map fills by state, so
                            the state is the unit that makes both work; the Cities
                            page reads the day-by-day cities back out.

Writes src/js/02-seed.js (baked into index.html by build.py) and data/trip-data.js
(the file the app reads on a fresh clone). Run it again if you change a sheet.
"""
import hashlib
import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
WORLD_BOOK = os.path.join(HERE, "World_tour_budget.xlsx")
INDIA_BOOK = os.path.join(HERE, "India_tour_budget.xlsx")

# Column letter in the sheet -> category id in the app.
COLS = {
    "I": "flights", "J": "transport", "K": "lodging", "L": "food",
    "M": "activities", "N": "carRental", "O": "visa", "P": "misc", "Q": "misc",
}
COLNUM = {"I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14, "O": 15, "P": 16, "Q": 17}

# Sheet's country label -> (clean name, iso2, default city, IATA, currency, lat, lon).
# The coordinates put a dot on the world map; the polygon fill comes from iso2.
# The sheet had three typos and filed Copenhagen under the country column.
META = {
    "Nepal":           ("Nepal", "NP", "Kathmandu", "KTM", "NPR", 27.72, 85.32),
    "Bhutan":          ("Bhutan", "BT", "Thimphu", "PBH", "BTN", 27.47, 89.64),
    "Vietnam":         ("Vietnam", "VN", "Hanoi", "HAN", "VND", 21.03, 105.85),
    "Sri Lanka":       ("Sri Lanka", "LK", "Colombo", "CMB", "LKR", 6.93, 79.86),
    "Indonisia":       ("Indonesia", "ID", "Bali", "DPS", "IDR", -8.65, 115.22),
    "Thailand":        ("Thailand", "TH", "Bangkok", "BKK", "THB", 13.76, 100.5),
    "Cambodia":        ("Cambodia", "KH", "Siem Reap", "REP", "KHR", 13.36, 103.86),
    "Laos":            ("Laos", "LA", "Luang Prabang", "LPQ", "LAK", 19.89, 102.14),
    "Malaysia":        ("Malaysia", "MY", "Kuala Lumpur", "KUL", "MYR", 3.14, 101.69),
    "Phillipines":     ("Philippines", "PH", "Manila", "MNL", "PHP", 14.6, 120.98),
    "Egypt":           ("Egypt", "EG", "Cairo", "CAI", "EGP", 30.04, 31.24),
    "Jordan":          ("Jordan", "JO", "Amman", "AMM", "JOD", 31.95, 35.93),
    "Czech Republic":  ("Czech Republic", "CZ", "Prague", "PRG", "CZK", 50.08, 14.44),
    "China":           ("China", "CN", "Beijing", "PEK", "CNY", 39.9, 116.41),
    "Singapore":       ("Singapore", "SG", "Singapore", "SIN", "SGD", 1.35, 103.82),
    "Japan":           ("Japan", "JP", "Tokyo", "TYO", "JPY", 35.68, 139.69),
    "Hungary":         ("Hungary", "HU", "Budapest", "BUD", "HUF", 47.5, 19.04),
    "Croatia":         ("Croatia", "HR", "Dubrovnik", "DBV", "EUR", 42.65, 18.09),
    "Germany":         ("Germany", "DE", "Berlin", "BER", "EUR", 52.52, 13.4),
    "Copenhagen":      ("Denmark", "DK", "Copenhagen", "CPH", "DKK", 55.68, 12.57),
    "Faroe Islands":   ("Faroe Islands", "FO", "Torshavn", "FAE", "DKK", 62.01, -6.77),
    "Turkey":          ("Turkey", "TR", "Istanbul", "IST", "TRY", 41.01, 28.98),
    "Georgia":         ("Georgia", "GE", "Tbilisi", "TBS", "GEL", 41.72, 44.79),
    "South Africa":    ("South Africa", "ZA", "Cape Town", "CPT", "ZAR", -33.92, 18.42),
    "Columbia":        ("Colombia", "CO", "Bogota", "BOG", "COP", 4.71, -74.07),
    "Peru":            ("Peru", "PE", "Cusco", "CUZ", "PEN", -13.53, -71.97),
    "Ecuador":         ("Ecuador", "EC", "Quito", "UIO", "USD", -0.18, -78.47),
    "Bolivia":         ("Bolivia", "BO", "La Paz", "LPB", "BOB", -16.5, -68.15),
    "Guatemala":       ("Guatemala", "GT", "Antigua Guatemala", "GUA", "GTQ", 14.56, -90.73),
    "Nicaragua":       ("Nicaragua", "NI", "Granada", "MGA", "NIO", 11.93, -85.96),
    "Mexico":          ("Mexico", "MX", "Mexico City", "MEX", "MXN", 19.43, -99.13),
    "Kyrgyzstan":      ("Kyrgyzstan", "KG", "Bishkek", "FRU", "KGS", 42.87, 74.59),
    "Uzbekistan":      ("Uzbekistan", "UZ", "Tashkent", "TAS", "UZS", 41.3, 69.24),
    "Tajikistan":      ("Tajikistan", "TJ", "Dushanbe", "DYU", "TJS", 38.56, 68.79),
    "Kazakhstan":      ("Kazakhstan", "KZ", "Almaty", "ALA", "KZT", 43.24, 76.89),
    "Armenia":         ("Armenia", "AM", "Yerevan", "EVN", "AMD", 40.18, 44.51),
}

# code -> (display name, units per 1 INR). Rates come from the sheet's A/B columns.
# EUR is the one correction: the sheet's Euro cells pointed at the GBP cell.
CURRENCIES = {
    "INR": ("Indian Rupee", 1.0),
    "NPR": ("Nepalese Rupee", 1.6),
    "BTN": ("Ngultrum", 1.0),
    "VND": ("Vietnamese Dong", 273.19),
    "LKR": ("Sri Lankan Rupee", 3.43),
    "IDR": ("Indonesian Rupiah", 185.67),
    "THB": ("Thai Baht", 0.35),
    "KHR": ("Cambodian Riel", 42.42),
    "LAK": ("Lao Kip", 235.0),
    "MYR": ("Malaysian Ringgit", 0.042),
    "PHP": ("Philippine Peso", 0.65),
    "EGP": ("Egyptian Pound", 0.532),
    "JOD": ("Jordanian Dinar", 0.0074),
    "CZK": ("Czech Koruna", 0.218),
    "CNY": ("Renminbi", 0.07),
    "SGD": ("Singapore Dollar", 0.013),
    "JPY": ("Japanese Yen", 1.67),
    "HUF": ("Hungarian Forint", 3.3),
    "EUR": ("Euro", 0.00901),
    "DKK": ("Danish Krone", 0.0674),
    "TRY": ("Turkish Lira", 0.505),
    "GEL": ("Georgian Lari", 0.028),
    "ZAR": ("South African Rand", 0.169),
    "COP": ("Colombian Peso", 33.39),
    "PEN": ("Peruvian Sol", 0.035),
    "USD": ("US Dollar", 1 / 95.44),
    "BOB": ("Boliviano", 0.072),
    "GTQ": ("Guatemalan Quetzal", 0.081),
    "NIO": ("Nicaraguan Cordoba", 0.383),
    "MXN": ("Mexican Peso", 0.178),
    "KGS": ("Kyrgyzstani Som", 0.916),
    "UZS": ("Uzbekistani Som", 124.5),
    "TJS": ("Tajikistani Somoni", 0.114),
    "KZT": ("Kazakhstani Tenge", 4.86),
    "AMD": ("Armenian Dram", 4.03),
    "GBP": ("Pound Sterling", 1 / 129.6),
    "AED": ("UAE Dirham", 0.0385),
}

CATEGORIES = [
    ("flights",    "Flights"),
    ("transport",  "Transportation"),
    ("lodging",    "Lodging"),
    ("food",       "Food"),
    ("activities", "Activities"),
    ("carRental",  "Car Rental"),
    ("visa",       "Visa"),
    ("misc",       "Misc"),
]


def read_world_places():
    wb = openpyxl.load_workbook(WORLD_BOOK, data_only=True)
    ws = wb["World Tour Budget"]
    places, order = [], 0
    for row in range(6, ws.max_row + 1):
        label = ws.cell(row=row, column=3).value
        if not label or not str(label).strip():
            continue
        label = str(label).strip()
        if label not in META:
            print("  ! no metadata for %r, skipping" % label)
            continue
        name, iso2, city, iata, cur, lat, lon = META[label]
        sheet_city = ws.cell(row=row, column=4).value
        if sheet_city and str(sheet_city).strip():
            city = str(sheet_city).strip()
        budget = {cid: 0 for cid, _ in CATEGORIES}
        for letter, cid in COLS.items():
            v = ws.cell(row=row, column=COLNUM[letter]).value
            if isinstance(v, (int, float)):
                budget[cid] += int(round(v))
        order += 1
        places.append({
            "id": "p_" + iso2.lower(),
            "name": name, "country": name, "city": city,
            "iso2": iso2, "iata": iata, "currency": cur,
            "lat": lat, "lon": lon,
            "days": 10, "order": order, "notes": "",
            "budget": budget, "images": [], "driveLinks": [], "itinerary": [],
        })
    return places


# --------------------------------------------------------------- india

# Column letter in the India sheet -> category id.
INDIA_COLS = {"I": "railway", "J": "intercity", "K": "lodging", "L": "food",
              "M": "activities", "N": "petrol", "O": "misc"}
INDIA_COLNUM = {"I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14, "O": 15}

INDIA_CATEGORIES = [
    ("railway",    "Flights & Rail"),
    ("intercity",  "Intercity Travel"),
    ("lodging",    "Lodging"),
    ("food",       "Food"),
    ("activities", "Activities"),
    ("petrol",     "Petrol"),
    ("misc",       "Misc"),
]

# State -> (gateway city, IATA, lat, lon). The coordinates drop the dot on the
# map; the state name itself is what fills the shape.
STATE_META = {
    "Punjab":            ("Amritsar", "ATQ", 31.63, 74.87),
    "Haryana":           ("Kurukshetra", "DEL", 29.97, 76.88),
    "Himachal Pradesh":  ("Shimla", "SLV", 31.10, 77.17),
    "Uttarakhand":       ("Haridwar", "DED", 29.95, 78.16),
    "Rajasthan":         ("Jaipur", "JAI", 26.92, 75.79),
    "Gujarat":           ("Ahmedabad", "AMD", 23.02, 72.57),
    "Madhya Pradesh":    ("Bhopal", "BHO", 23.26, 77.41),
    "Uttar Pradesh":     ("Agra", "AGR", 27.18, 78.02),
    "Bihar":             ("Patna", "PAT", 25.59, 85.14),
    "Jharkhand":         ("Ranchi", "IXR", 23.34, 85.31),
    "West Bengal":       ("Kolkata", "CCU", 22.57, 88.36),
    "Sikkim":            ("Gangtok", "IXB", 27.33, 88.61),
    "Assam":             ("Guwahati", "GAU", 26.14, 91.74),
    "Arunachal Pradesh": ("Tawang", "TEZ", 27.59, 91.87),
    "Meghalaya":         ("Shillong", "SHL", 25.58, 91.89),
    "Manipur":           ("Imphal", "IMF", 24.82, 93.94),
    "Nagaland":          ("Kohima", "DMU", 25.67, 94.11),
    "Mizoram":           ("Aizawl", "AJL", 23.73, 92.72),
    "Tripura":           ("Agartala", "IXA", 23.83, 91.28),
    "Odisha":            ("Bhubaneswar", "BBI", 20.30, 85.82),
    "Chhattisgarh":      ("Raipur", "RPR", 21.25, 81.63),
    "Telangana":         ("Hyderabad", "HYD", 17.39, 78.49),
    "Andhra Pradesh":    ("Visakhapatnam", "VTZ", 17.69, 83.22),
    "Karnataka":         ("Bangalore", "BLR", 12.97, 77.59),
    "Kerala":            ("Kochi", "COK", 9.93, 76.27),
    "Tamil Nadu":        ("Chennai", "MAA", 13.08, 80.27),
    "Goa":               ("Panaji", "GOI", 15.50, 73.83),
    "Maharashtra":       ("Mumbai", "BOM", 19.08, 72.88),
}


def read_india_places():
    """One place per state. Each day row becomes an itinerary entry so the
    cities, and what you meant to do in them, survive the import."""
    wb = openpyxl.load_workbook(INDIA_BOOK, data_only=True)
    ws = wb["India Tour"]
    places, order, current = [], 0, None

    for row in range(6, 300):
        state = ws.cell(row=row, column=3).value
        day = ws.cell(row=row, column=2).value
        city = ws.cell(row=row, column=4).value
        note = ws.cell(row=row, column=5).value

        if state and str(state).strip():
            state = str(state).strip()
            if state not in STATE_META:
                print("  ! no metadata for state %r, skipping" % state)
                current = None
                continue
            gateway, iata, lat, lon = STATE_META[state]
            budget = {cid: 0 for cid, _ in INDIA_CATEGORIES}
            for letter, cid in INDIA_COLS.items():
                v = ws.cell(row=row, column=INDIA_COLNUM[letter]).value
                if isinstance(v, (int, float)):
                    budget[cid] += int(round(v))
            order += 1
            current = {
                "id": "in_" + slugify(state),
                "name": state, "country": state, "city": gateway,
                "iso2": "IN", "iata": iata, "currency": "INR",
                "lat": lat, "lon": lon,
                "days": 0, "order": order, "notes": "",
                "budget": budget, "images": [], "driveLinks": [], "itinerary": [],
            }
            places.append(current)
            continue

        if current is None or not city:
            continue

        # "Day 1" on the first row, then bare numbers
        n = len(current["itinerary"]) + 1
        if isinstance(day, (int, float)):
            n = int(day)
        elif isinstance(day, str):
            digits = "".join(ch for ch in day if ch.isdigit())
            if digits:
                n = int(digits)
        current["itinerary"].append({
            "id": "it_%s_%d" % (current["id"], n),
            "day": n,
            "title": str(city).strip(),
            "note": str(note).strip() if note else "",
            "cost": 0,
            "done": False,
        })
        current["days"] = max(current["days"], n)

    for p in places:
        p["days"] = max(1, p["days"])
    return places


def slugify(v):
    return "".join(ch if ch.isalnum() else "-" for ch in v.lower()).strip("-")


def build():
    places = read_world_places()
    total = sum(sum(p["budget"].values()) for p in places)
    print("  world : %d countries, INR %s" % (len(places), format(total, ",")))

    india = read_india_places()
    india_total = sum(sum(p["budget"].values()) for p in india)
    india_days = sum(p["days"] for p in india)
    india_cities = sum(len(p["itinerary"]) for p in india)
    print("  india : %d states, %d days, %d city-days, INR %s"
          % (len(india), india_days, india_cities, format(india_total, ",")))

    seed = {
        "version": 1,
        "settings": {
            "name": "",
            "symbol": "₹",
            "homeAirport": "DEL",
            "homeCity": "Delhi",
            "adults": 1,
            "linkTemplates": {
                "flights": {
                    "label": "Skyscanner",
                    "url": "https://www.skyscanner.co.in/transport/flights/{home}/{iata}/{yymmdd}/",
                },
                "stays": {
                    "label": "Booking.com",
                    "url": "https://www.booking.com/searchresults.html?ss={city}%2C+{country}&checkin={in}&checkout={out}&group_adults={adults}",
                },
                "activities": {
                    "label": "GetYourGuide",
                    "url": "https://www.getyourguide.com/s/?q={city}",
                },
            },
        },
        "currencies": {c: {"name": n, "perINR": r} for c, (n, r) in CURRENCIES.items()},
        "activeTrip": "trip_world",
        "trips": [
            {
                "id": "trip_world",
                "name": "World Tour",
                "kind": "world",
                "start": "", "end": "",
                "note": "Imported from World_tour_budget.xlsx",
                "defaultDays": 10,
                "categories": [{"id": c, "label": l} for c, l in CATEGORIES],
                "places": places,
                "expenses": [],
            },
            {
                "id": "trip_india",
                "name": "Within India",
                "kind": "domestic",
                "start": "", "end": "",
                "note": "Imported from India_tour_budget.xlsx",
                "defaultDays": 6,
                "categories": [{"id": c, "label": l} for c, l in INDIA_CATEGORIES],
                "places": india,
                "expenses": [],
            },
        ],
    }

    # A fingerprint of what the sheets produced. The app compares it with the
    # copy in the browser so a re-import never sits invisible behind stale data.
    seed["stamp"] = hashlib.sha1(
        json.dumps(seed["trips"], ensure_ascii=False, sort_keys=True).encode()
    ).hexdigest()[:12]
    seed["stampNote"] = "%d countries, %d states, %s days" % (
        len(places), len(india), sum(p["days"] for p in places + india))

    blob = json.dumps(seed, ensure_ascii=False, indent=1)
    with open(os.path.join(HERE, "src", "js", "02-seed.js"), "w") as fh:
        fh.write("/* Generated by import_xlsx.py from World_tour_budget.xlsx. Do not hand-edit. */\n")
        fh.write("window.DUNIYA_SEED = " + blob + ";\n")
    with open(os.path.join(HERE, "data", "trip-data.js"), "w") as fh:
        fh.write("/* Your trips. Written by the app; safe to commit, diff and pull. */\n")
        fh.write("window.DUNIYA_DATA = " + blob + ";\n")
    print("  wrote src/js/02-seed.js and data/trip-data.js")


if __name__ == "__main__":
    build()
